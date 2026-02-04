import flet as ft
import json
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
# ================= MODELO / LÓGICA =================

class CostManager:
    def __init__(self, filename="gastos.xlsx", dict_file="costos.json"):
        self.filename = filename
        self.dict_file = dict_file
        self.expenses = []
        self.cost_dict = {}

        self.load_cost_dict()
        self.load_expenses()

    def load_cost_dict(self):
        if os.path.exists(self.dict_file):
            try:
                with open(self.dict_file, 'r', encoding='utf-8') as f:
                    self.cost_dict = json.load(f)
            except Exception as e:
                print(f"Error cargando costos: {e}")
                self.cost_dict = {}
        else:
            # Default Data
            self.cost_dict = {
                "Pescado (Kg)": 18.0,
                "Limón (Kg)": 7.0,
                "Cebolla (Kg)": 3.5,
                "Mesero (Día)": 50.0,
                "Aceite (L)": 8.5
            }
            self.save_cost_dict()

    def save_cost_dict(self):
        try:
            with open(self.dict_file, 'w', encoding='utf-8') as f:
                json.dump(self.cost_dict, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"Error guardando costos: {e}")

    def add_cost_item(self, name, cost):
        self.cost_dict[name] = float(cost)
        self.save_cost_dict()

    def delete_cost_item(self, name):
        if name in self.cost_dict:
            del self.cost_dict[name]
            self.save_cost_dict()

    def get_next_id(self):
        if not self.expenses:
            return 1
        return max(e['id'] for e in self.expenses) + 1

    def load_expenses(self):
        if not os.path.exists(self.filename):
            return

        try:
            wb = load_workbook(self.filename)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None: continue
                try:
                    expense = {
                        'id': int(row[0]),
                        'fecha': row[1],
                        'item': row[2],
                        'cantidad': float(row[3]),
                        'precio_unit': float(row[4]),
                        'total': float(row[5])
                    }
                    self.expenses.append(expense)
                except Exception:
                    pass
            # Sort by Date Descending
            self.expenses.sort(key=lambda x: x['fecha'], reverse=True)
        except Exception as e:
            print(f"Error cargando gastos: {e}")

    def save_expenses(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial Gastos"
        headers = ["ID", "Fecha", "Insumo", "Cantidad", "Costo Unit.", "Total"]
        ws.append(headers)
        
        for e in self.expenses:
            ws.append([
                e['id'], e['fecha'], e['item'], e['cantidad'], e['precio_unit'], e['total']
            ])
        try:
            wb.save(self.filename)
        except PermissionError as e:
            return str(e)
        return None

    def add_expense(self, item, cantidad, date_str=None):
        if item not in self.cost_dict: return "Item no existe"
        
        cost = self.cost_dict[item]
        if not date_str:
            date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
        expense = {
            'id': self.get_next_id(),
            'fecha': date_str,
            'item': item,
            'cantidad': cantidad,
            'precio_unit': cost,
            'total': cost * cantidad
        }
        self.expenses.insert(0, expense) # Add to top
        # Sort again just in case date was in past
        self.expenses.sort(key=lambda x: x['fecha'], reverse=True)
        return self.save_expenses()

    def delete_expense(self, exp_id):
        self.expenses = [e for e in self.expenses if e['id'] != exp_id]
        return self.save_expenses()

    def update_expense_date(self, exp_id, new_date):
        for e in self.expenses:
            if e['id'] == exp_id:
                # Keep time if only date is gathered? Or expect full datetime iso string?
                # User picker returns YYYY-MM-DD. We might want to keep time or just set time to 00:00.
                # Simplification: Append current time if input is only date? 
                # Or just replace string.
                e['fecha'] = new_date
                self.expenses.sort(key=lambda x: x['fecha'], reverse=True)
                return self.save_expenses()
        return None

    def get_financials(self, start_date=None, end_date=None):
        df_exp = pd.DataFrame(self.expenses)
        
        total_expenses = 0
        daily_expenses = {}
        
        if not df_exp.empty:
            try:
                df_exp['fecha_dt'] = pd.to_datetime(df_exp['fecha'])
                if start_date and end_date:
                    mask = (df_exp['fecha_dt'] >= pd.to_datetime(start_date)) & (df_exp['fecha_dt'] <= pd.to_datetime(end_date) + pd.Timedelta(days=1))
                    df_filtered = df_exp.loc[mask]
                else:
                    df_filtered = df_exp # All if no filter? Or match logic of Sales?
                    # For financials, usually we want Total if no filter or Today?
                    # Let's default to All Time if no filter for Utility, or Month?
                    # User request: "Dashboard BI Financiero... filtrable". 
                    pass
                
                if start_date and end_date:
                     total_expenses = df_filtered['total'].sum()
                     daily_expenses = df_filtered.groupby(df_filtered['fecha_dt'].dt.date)['total'].sum().to_dict()
                else:
                    # If no date, maybe return 0 or Total? Let's return Total for now but filtered by today in dashboard default logic 
                    # Actually OrderManager defaults to TODAY. Let's make CostManager consistent or flexible.
                    # Let's handle logic in Dashboard.
                    pass

            except Exception:
                pass
        
        return total_expenses, daily_expenses
# ================= MODELO / LÓGICA =================

class OrderManager:
    def __init__(self, filename="pedidos_cevicheria.xlsx", menu_file="menu.json"):
        self.filename = filename
        self.menu_file = menu_file
        self.orders = []
        self.menu = {}

        self.load_menu()
        self.load_orders()

    def load_menu(self):
        if os.path.exists(self.menu_file):
            try:
                with open(self.menu_file, 'r', encoding='utf-8') as f:
                    self.menu = json.load(f)
            except Exception as e:
                print(f"Error cargando menú: {e}")
                self.menu = {}
        else:
            self.menu = {
                "Duo Marino": 15.0,
                "Causa de Pescado": 10.0,
                "Ceviche": 12.0,
                "Trio Marino": 20.0,
            }
            self.save_menu()

    def save_menu(self):
        try:
            with open(self.menu_file, 'w', encoding='utf-8') as f:
                json.dump(self.menu, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"Error guardando menú: {e}")

    def add_dish(self, name, price):
        # Update if exists, else add new
        self.menu[name] = float(price)
        self.save_menu()

    def delete_dish(self, name):
        if name in self.menu:
            del self.menu[name]
            self.save_menu()

    def get_next_id(self):
        if not self.orders:
            return 1
        return max(o['id'] for o in self.orders) + 1

    def load_orders(self):
        if not os.path.exists(self.filename):
            return

        try:
            wb = load_workbook(self.filename)
            ws = wb.active
            self.orders = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None: continue
                
                row_data = list(row)
                while len(row_data) < 10:
                    row_data.append(None)

                try:
                    order = {
                        'id': int(row_data[0]),
                        'fecha': row_data[1],
                        'cliente': row_data[2],
                        'plato': row_data[3],
                        'cantidad': int(row_data[4]),
                        'precio': float(row_data[5]),
                        'subtotal': float(row_data[6]) if row_data[6] is not None else (int(row_data[4]) * float(row_data[5])),
                        'metodo_pago': str(row_data[7]) if row_data[7] else "Efectivo",
                        'entregado': str(row_data[8]) == 'Si',
                        'pagado': str(row_data[9]) == 'Si'
                    }
                    self.orders.append(order)
                except Exception:
                    pass
            # Sort Descending
            self.orders.sort(key=lambda x: x['fecha'], reverse=True)
        except Exception as e:
            print(f"Error cargando historial: {e}")

    def save_orders(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial Pedidos"
        headers = ["ID", "Fecha", "Cliente", "Plato", "Cant.", "Precio Unit.", "Total", "Método Pago", "Entregado", "Pagado"]
        ws.append(headers)
        
        for o in self.orders:
            ws.append([
                o['id'], o['fecha'], o['cliente'], o['plato'], o['cantidad'], o['precio'],
                o['subtotal'], o.get('metodo_pago', 'Efectivo'),
                "Si" if o['entregado'] else "No", "Si" if o['pagado'] else "No"
            ])
        try:
            wb.save(self.filename)
        except PermissionError as e:
            return str(e)
        return None

    def add_order(self, cliente, plato, cantidad, metodo_pago, date_str=None):
        if plato not in self.menu: return None
        precio = self.menu[plato]
        
        if not date_str:
            date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
        order = {
            'id': self.get_next_id(),
            'fecha': date_str,
            'cliente': cliente,
            'plato': plato,
            'cantidad': cantidad,
            'precio': precio,
            'subtotal': precio * cantidad,
            'metodo_pago': metodo_pago,
            'entregado': False,
            'pagado': False
        }
        self.orders.append(order)
        # Sort again just in case date was in past
        self.orders.sort(key=lambda x: x['fecha'], reverse=True)
        return self.save_orders()

    def delete_order(self, order_id):
        self.orders = [o for o in self.orders if o['id'] != order_id]
        return self.save_orders()

    def toggle_status(self, order_id, field):
        for order in self.orders:
            if order['id'] == order_id:
                order[field] = not order[field]
                return self.save_orders()
        return None

    def get_filtered_stats(self, start_date=None, end_date=None):
        if not self.orders:
            return None
            
    def update_order_date(self, order_id, new_date):
        for o in self.orders:
            if o['id'] == order_id:
                o['fecha'] = new_date
                self.orders.sort(key=lambda x: x['fecha'], reverse=True)
                return self.save_orders()
        return None

    def get_filtered_stats(self, start_date=None, end_date=None):
        if not self.orders:
            return None
            
        df = pd.DataFrame(self.orders)
        
        # Date Conversion
        try:
            df['fecha_dt'] = pd.to_datetime(df['fecha'])
        except Exception:
            return None

        # Filter by Date Range
        if start_date and end_date:
            mask = (df['fecha_dt'] >= pd.to_datetime(start_date)) & (df['fecha_dt'] <= pd.to_datetime(end_date) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1))
            df_filtered = df.loc[mask]
        else:
             # Default to today if no range
            today = datetime.now().strftime("%Y-%m-%d")
            df_filtered = df[df['fecha'].str.startswith(today)]

        if df_filtered.empty:
            return {
                "total_sales": 0,
                "ticket_average": 0,
                "top_3_dishes": [],
                "bottom_3_dishes": [],
                "top_3_clients": [],
                "avg_price_per_dish": 0,
                "payment_methods": {},
                "daily_sales_trend": {},
                "rush_hour": {h: 0 for h in range(24)}
            }

        # KPIs
        total_sales = df_filtered['subtotal'].sum()
        ticket_average = df_filtered['subtotal'].mean()
        
        # Top/Bottom Dishes
        dish_counts = df_filtered['plato'].value_counts()
        total_items = dish_counts.sum()
        
        top_3_dishes = [{"name": name, "pct": (count/total_items)*100} for name, count in dish_counts.head(3).items()]
        bottom_3_dishes = [{"name": name, "pct": (count/total_items)*100} for name, count in dish_counts.tail(3).items()]

        # Top Clients
        client_counts = df_filtered['cliente'].value_counts()
        total_clients = client_counts.sum()
        top_3_clients = [{"name": name, "pct": (count/total_clients)*100} for name, count in client_counts.head(3).items()]

        # Avg Price per Dish (Total Sales / Total Qty)
        total_qty = df_filtered['cantidad'].sum()
        avg_price_per_dish = total_sales / total_qty if total_qty > 0 else 0

        # Payment Methods
        payment_methods = df_filtered['metodo_pago'].value_counts().to_dict()

        # Daily Sales Trend (for LineChart)
        daily_sales = df_filtered.groupby(df_filtered['fecha_dt'].dt.date)['subtotal'].sum().to_dict()
        daily_sales = dict(sorted(daily_sales.items()))
        
        # Rush Hour (Orders per Hour)
        df_filtered['hour'] = df_filtered['fecha_dt'].dt.hour
        hourly_counts = df_filtered['hour'].value_counts().sort_index().to_dict()
        rush_hour = {h: hourly_counts.get(h, 0) for h in range(24)}

        return {
            "total_sales": total_sales,
            "ticket_average": ticket_average,
            "top_3_dishes": top_3_dishes,
            "bottom_3_dishes": bottom_3_dishes,
            "top_3_clients": top_3_clients,
            "avg_price_per_dish": avg_price_per_dish,
            "payment_methods": payment_methods,
            "daily_sales_trend": daily_sales,
            "rush_hour": rush_hour
        }

# ================= VISTA / UI (FLET) =================

def main(page: ft.Page):
    page.title = "Cevichería YAFRANK"
    page.theme_mode = ft.ThemeMode.LIGHT
    page.padding = 0
    page.window.min_width = 1000
    page.window.min_height = 700

    manager = OrderManager()
    cost_manager = CostManager()
    
    # 1. SALES VIEW COMPONENT
    def create_sales_view():
        
        def hover_effect(e):
            e.control.scale = 1.02 if e.data == "true" else 1.0
            e.control.bgcolor = ft.Colors.SURFACE_CONTAINER_HIGHEST if e.data == "true" else ft.Colors.SURFACE
            e.control.update()

        menu_items_container = ft.Column(scroll=ft.ScrollMode.AUTO, expand=True, spacing=10)
        
        qty_input = ft.TextField(value="1", label="Cantidad", width=80, text_align="center", keyboard_type="number")
        
        payment_group = ft.RadioGroup(content=ft.Row([
            ft.Radio(value="Efectivo", label="Efectivo"),
            ft.Radio(value="Yape", label="Yape"),
            ft.Radio(value="Plin", label="Plin"),
        ]))
        payment_group.value = "Efectivo"
        
        search_input = ft.TextField(
            label="Buscar Cliente/Plato", 
            prefix_icon=ft.Icons.SEARCH,
            on_change=lambda e: filter_orders(e.control.value),
            expand=True
        )
        
        # Date Picker for New Orders
        order_date_picker = ft.DatePicker(
            first_date=datetime(2020, 1, 1),
            last_date=datetime(2100, 12, 31)
        )
        page.overlay.append(order_date_picker)
        
        client_input = ft.TextField(label="Nombre Cliente", expand=True)
        date_btn = ft.IconButton(icon=ft.Icons.CALENDAR_MONTH, tooltip="Fecha del Pedido", on_click=lambda _: page.open(order_date_picker))

        # --- TABLA DE VENTAS CON HEADER FIJO (Sticky Header) ---
        col_widths = [50, 120, 100, 150, 60, 80, 80, 100, 120]
        headers_txt = ["ID", "Fecha", "Cliente", "Plato", "Cant.", "Total", "Pago", "Estado", "Acciones"]
        
        header_row = ft.Row(
            controls=[
                ft.Container(content=ft.Text(h, weight="bold"), width=w) 
                for h, w in zip(headers_txt, col_widths)
            ],
            spacing=10
        )
        
        orders_header = ft.Container(
            content=header_row,
            bgcolor=ft.Colors.SURFACE_CONTAINER_HIGHEST,
            padding=10,
            border_radius=ft.border_radius.only(top_left=10, top_right=10),
        )

        orders_list = ft.ListView(expand=True, spacing=0)

        # Helper for Date Editing
        def edit_date_click(e, order_id, current_date):
            def save_date(e2):
                if date_picker.value:
                    # Keep time? User request just Date Picker.
                    # If we lose time, order order might change if we use time for sort.
                    # Append current time or 00:00?
                    # Or just YYYY-MM-DD.
                    new_d = date_picker.value.strftime("%Y-%m-%d")
                    # Try to keep original time part of current_date string if compatible
                    final_d = f"{new_d} {current_date[11:]}" if len(current_date) > 10 else new_d
                
                    manager.update_order_date(order_id, final_d)
                refresh_orders_table_logic()
                page.close(dlg)
                page.update()

            date_picker = ft.DatePicker(
                first_date=datetime(2020, 1, 1),
                last_date=datetime(2100, 12, 31)
            )
            page.overlay.append(date_picker)
        
            dlg = ft.AlertDialog(
                title=ft.Text("Editar Fecha"),
                content=ft.Column([
                    ft.Text(f"Fecha actual: {current_date}"),
                    ft.ElevatedButton("Seleccionar Nueva Fecha", icon=ft.Icons.CALENDAR_MONTH, on_click=lambda _: page.open(date_picker))
                ], height=100),
                actions=[
                    ft.TextButton("Cancelar", on_click=lambda _: page.close(dlg)),
                    ft.TextButton("Guardar", on_click=save_date)
                ],
            )
            page.open(dlg)

        def refresh_orders_table_logic(orders_to_show=None):
            orders_list.controls.clear()
            data_source = orders_to_show if orders_to_show is not None else manager.orders
            # Already sorted by manager load/add, but ensure
            sorted_orders = sorted(data_source, key=lambda x: x['id'], reverse=True) # Sort by ID desc usually matches Date desc roughly for recent
            # Better sort by Date
            sorted_orders.sort(key=lambda x: x['fecha'], reverse=True)

            for o in sorted_orders[:50]:
                status_paid = "Pagado" if o['pagado'] else "Pendiente"
                color_paid = ft.Colors.GREEN if o['pagado'] else ft.Colors.RED
                status_del = "Entregado" if o['entregado'] else "Cocina"
                color_del = ft.Colors.BLUE if o['entregado'] else ft.Colors.ORANGE
            
                # Action Buttons (unchanged)
                actions = ft.Row([
                     ft.IconButton("check", icon_color=ft.Colors.GREEN, 
                        on_click=lambda e, oid=o['id']: (manager.toggle_status(oid, 'entregado'), refresh_orders_table_logic(), page.update())),
                     ft.IconButton("attach_money", icon_color=ft.Colors.BLUE,
                        on_click=lambda e, oid=o['id']: (manager.toggle_status(oid, 'pagado'), refresh_orders_table_logic(), page.update())),
                     ft.IconButton("delete", icon_color=ft.Colors.RED,
                        on_click=lambda e, oid=o['id']: (manager.delete_order(oid), refresh_orders_table_logic(), page.update()))
                ])

                row_controls = [
                    ft.Text(str(o['id'])),
                    # Date Button for Edit
                    ft.TextButton(
                        str(o['fecha'])[:16], 
                        on_click=lambda e, oid=o['id'], cd=o['fecha']: edit_date_click(e, oid, cd)
                    ),
                    ft.Text(o['cliente']),
                    ft.Text(o['plato']),
                    ft.Text(str(o['cantidad'])),
                    ft.Text(f"S/{o['subtotal']:.2f}"),
                    ft.Text(o['metodo_pago']),
                    ft.Row([
                        ft.Container(content=ft.Text(status_paid, size=10, color="white"), bgcolor=color_paid, padding=5, border_radius=5),
                        ft.Container(content=ft.Text(status_del, size=10, color="white"), bgcolor=color_del, padding=5, border_radius=5)
                    ]),
                    actions
                ]
            
                cells = [ft.Container(content=c, width=w) for c, w in zip(row_controls, col_widths)]
            
                orders_list.controls.append(
                    ft.Container(
                        content=ft.Row(cells, spacing=10),
                        padding=ft.padding.symmetric(vertical=5, horizontal=10),
                        border=ft.border.only(bottom=ft.border.BorderSide(1, ft.Colors.GREY_200)),
                        bgcolor=ft.Colors.SURFACE,
                        scale=1.0,
                        animate_scale=ft.animation.Animation(300, ft.AnimationCurve.EASE_OUT),
                        on_hover=hover_effect
                    )
                )

        def refresh_menu_logic():
             # PURE LOGIC: Modifies the Control's state but DOES NOT call .update()
            menu_items_container.controls.clear()
            for dish, price in manager.menu.items():
                card = ft.Container(
                    content=ft.Row([
                        # Re-design: Price inline with SpaceBetween
                        ft.Row([
                             ft.Text(dish, weight="bold", size=16, color=ft.Colors.ON_SURFACE, expand=True),
                             ft.Text(f"S/ {price:.2f}", color=ft.Colors.PRIMARY, size=16, weight="bold")
                        ], expand=True, alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment="center"),
                        
                        ft.IconButton(
                            icon="add_circle", 
                            icon_color=ft.Colors.PRIMARY, 
                            icon_size=30,
                            tooltip="Agregar Pedido",
                            on_click=lambda e, d=dish: add_order_click(e, d)
                        )
                    ], alignment="spaceBetween"),
                    padding=10,
                    bgcolor=ft.Colors.SURFACE,
                    border_radius=10,
                    scale=1.0,
                    animate_scale=ft.animation.Animation(300, ft.AnimationCurve.EASE_OUT),
                    on_hover=hover_effect
                )
                menu_items_container.controls.append(card)

        # Interaction Handlers
        def add_order_click(e, plato_name):
            if not client_input.value:
                # Assuming page.snack_bar is handled by page.update()
                snack = ft.SnackBar(ft.Text("Ingrese Nombre del Cliente"), bgcolor=ft.Colors.RED)
                page.overlay.append(snack)
                page.open(snack)
                page.update()
                return

            try:
                qty = int(qty_input.value)
            except ValueError:
                qty = 1
            
            # Get Date
            d_str = None
            if order_date_picker.value:
                d_str = order_date_picker.value.strftime("%Y-%m-%d %H:%M:%S")

            err = manager.add_order(client_input.value, plato_name, qty, payment_group.value, date_str=d_str)
            if err:
                snack = ft.SnackBar(ft.Text(f"Error guardando: {err}"), bgcolor=ft.Colors.RED)
                page.overlay.append(snack)
                page.update()
                page.open(snack)
            else:
                snack = ft.SnackBar(ft.Text(f"Pedido Agregado: {plato_name}"), bgcolor=ft.Colors.GREEN)
                page.overlay.append(snack)
                # Logic update
                refresh_orders_table_logic()
                if hasattr(create_dashboard_view, 'update_logic'):
                     create_dashboard_view.update_logic()
                
                page.update()
                page.open(snack)
            
            # page.update() handled above inside else/if blocks or just one at end?
            # User requirement: "Asegura que tras agregar un pedido, se llame a refresh_orders_table_logic() y se realice un page.update() inmediato."
            # My logic above does that. I will add a general page.update() at the end to be safe.
            page.update()

        def delete_order_click(e, oid):
            manager.delete_order(oid)
            refresh_orders_table_logic()
            page.update() # Update page

        def toggle_paid_click(e, oid):
            manager.toggle_status(oid, 'pagado')
            refresh_orders_table_logic()
            page.update() # Update page

        def toggle_delivered_click(e, oid):
            manager.toggle_status(oid, 'entregado')
            refresh_orders_table_logic()
            page.update() # Update page
        
        # Search Logic
        def filter_orders(query):
            if not query:
                refresh_orders_table_logic()
            else:
                q = query.lower()
                filtered = [o for o in manager.orders if q in o['cliente'].lower() or q in o['plato'].lower()]
                refresh_orders_table_logic(filtered)
            page.update()

        # --- INITIAL DATA POPULATION ---
        refresh_menu_logic() 
        refresh_orders_table_logic()
        
        # Expose methods for external access
        create_sales_view.refresh_table = refresh_orders_table_logic
        create_sales_view.refresh_menu = refresh_menu_logic

        view = ft.Row([
            # Left Column (Menu)
            ft.Container(
                content=ft.Column([
                    ft.Text("Carta", size=20, weight="bold"),
                    ft.Divider(),
                    ft.Row([client_input, date_btn]),
                    ft.Row([
                        qty_input,
                        payment_group
                    ], alignment="spaceBetween", vertical_alignment="center"),
                    ft.Divider(),
                    menu_items_container
                ]),
                width=350,
                padding=10,
                bgcolor=ft.Colors.SURFACE,
                border_radius=12,
            ),
            # Right Column (Orders)
            ft.Container(
                content=ft.Column([
                   ft.Row([
                       ft.Text("Pedidos Recientes", size=20, weight="bold"),
                       ft.IconButton("refresh", on_click=lambda e: (filter_orders(search_input.value), page.update()))
                   ], alignment="spaceBetween"),
                   ft.Row([search_input]), # Add Search Bar Row
                   ft.Row(
                       [
                           ft.Column([
                                orders_header,
                                orders_list
                           ], 
                           expand=True, 
                           spacing=0,
                           width=sum(col_widths) + 120 # Width sum + padding
                           )
                       ], 
                       scroll=ft.ScrollMode.AUTO, 
                       expand=True
                   ) 
                ]),
                expand=True,
                padding=10,
                border_radius=12,
                bgcolor=ft.Colors.SURFACE,
            )
        ], expand=True, spacing=10)
        
        return view

    # 2. DASHBOARD VIEW (Updated for Financial BI)
    def create_dashboard_view():
        # Date Pickers (Reuse existing logic pattern but inside function to capture closure)
        # Note: If we move this code, we need to ensure references match.
        
        start_date_picker = ft.DatePicker(
            on_change=lambda e: update_dashboard_logic(),
            first_date=datetime(2020, 1, 1),
            last_date=datetime(2100, 12, 31)
        )
        end_date_picker = ft.DatePicker(
             on_change=lambda e: update_dashboard_logic(),
            first_date=datetime(2020, 1, 1),
            last_date=datetime(2100, 12, 31)
        )
        page.overlay.append(start_date_picker)
        page.overlay.append(end_date_picker)

        btn_start_date = ft.ElevatedButton(
            "Desde", 
            icon=ft.Icons.CALENDAR_MONTH, 
            on_click=lambda _: page.open(start_date_picker)
        )
        btn_end_date = ft.ElevatedButton(
            "Hasta", 
            icon=ft.Icons.CALENDAR_MONTH, 
            on_click=lambda _: page.open(end_date_picker)
        )
        
        def clear_filters():
             start_date_picker.value = None
             end_date_picker.value = None
             update_dashboard_logic() 

        date_range_row = ft.Row([
            ft.Text("Filtrar por Fecha:", weight="bold"),
            btn_start_date,
            btn_end_date,
            ft.IconButton(icon=ft.Icons.FILTER_LIST_OFF, tooltip="Limpiar Filtros", on_click=lambda e: clear_filters())
        ], alignment="center", spacing=20)
        
        # Financial KPIs
        stat_income = ft.Text("S/ 0.00", size=20, weight="bold")
        stat_expenses = ft.Text("S/ 0.00", size=20, weight="bold")
        stat_profit = ft.Text("S/ 0.00", size=25, weight="bold", color=ft.Colors.GREEN)
        
        # Charts
        chart_payment = ft.PieChart(sections=[], sections_space=0, center_space_radius=40, expand=True)
        
        chart_financial = ft.BarChart(
            bar_groups=[],
            border=ft.border.all(1, ft.Colors.GREY_400),
            left_axis=ft.ChartAxis(labels_size=40, title=ft.Text("Monto S/", size=12)),
            bottom_axis=ft.ChartAxis(
                labels=[
                    ft.ChartAxisLabel(value=0, label=ft.Text("Ingresos")),
                    ft.ChartAxisLabel(value=1, label=ft.Text("Egresos")),
                ],
                labels_size=40,
            ),
            horizontal_grid_lines=ft.ChartGridLines(color=ft.Colors.GREY_300, width=1, dash_pattern=[3, 3]),
            tooltip_bgcolor=ft.Colors.with_opacity(0.8, ft.Colors.GREY_900),
            max_y=1000, # Dynamic
            expand=True
        )

        chart_rush_hour = ft.LineChart(
            data_series=[],
            border=ft.border.all(1, ft.Colors.GREY_400),
            left_axis=ft.ChartAxis(labels_size=40, title=ft.Text("Pedidos", size=12)),
            bottom_axis=ft.ChartAxis(
                labels=[ft.ChartAxisLabel(value=h, label=ft.Text(str(h))) for h in range(0, 24, 2)],
                labels_size=20,
            ),
            tooltip_bgcolor=ft.Colors.with_opacity(0.8, ft.Colors.GREY_900),
            min_y=0,
            expand=True
        )

        # Analysis containers
        top_dishes_col = ft.Column()
        bottom_dishes_col = ft.Column()
        top_clients_col = ft.Column()
        ai_insights_txt = ft.Text("", italic=True, size=14, color=ft.Colors.GREY_700)

        def generate_pdf(e):
            s_date = start_date_picker.value.strftime("%Y-%m-%d") if start_date_picker.value else "Inicio"
            e_date = end_date_picker.value.strftime("%Y-%m-%d") if end_date_picker.value else "Fin"
            filename = f"reporte_cierre_{s_date}_a_{e_date}.pdf".replace(" ", "_")
            
            try:
                c = canvas.Canvas(filename, pagesize=letter)
                width, height = letter
                
                # Header
                c.setFont("Helvetica-Bold", 18)
                c.drawString(50, height - 50, f"Cevichería YAFRANK - Reporte de Cierre")
                c.setFont("Helvetica", 12)
                c.drawString(50, height - 70, f"Periodo: {s_date} al {e_date}")
                c.line(50, height - 80, width - 50, height - 80)
                
                # Financials
                c.drawString(50, height - 110, f"Ingresos Totales: {stat_income.value}")
                c.drawString(50, height - 130, f"Egresos Totales: {stat_expenses.value}")
                c.setFont("Helvetica-Bold", 14)
                c.drawString(50, height - 160, f"Utilidad Neta: {stat_profit.value}")
                
                # --- Detail Sections ---
                y_pos = height - 200
                
                # Sales Detail
                c.setFont("Helvetica-Bold", 10)
                c.drawString(50, y_pos, "Detalle de Ventas")
                y_pos -= 20
                c.setFont("Helvetica", 8)
                # Ventas: ID | Fecha | Cliente | Plato | Cant. | Precio Plato | Total.
                # X: ID(30), Fecha(60), Cliente(140), Plato(260), Cant(380), Price(420), Total(480)
                c.drawString(30, y_pos, "ID")
                c.drawString(60, y_pos, "Fecha")
                c.drawString(140, y_pos, "Cliente")
                c.drawString(260, y_pos, "Plato")
                c.drawString(380, y_pos, "Cant.")
                c.drawString(420, y_pos, "P.Unit")
                c.drawString(480, y_pos, "Total")
                y_pos -= 15
                
                # Filter Sales
                sales_data = manager.orders
                if start_date_picker.value and end_date_picker.value:
                     s = pd.to_datetime(start_date_picker.value)
                     e = pd.to_datetime(end_date_picker.value) + pd.Timedelta(days=1)
                     sales_data = [o for o in sales_data if s <= pd.to_datetime(o['fecha']) <= e]

                for o in sales_data[:50]: # Expanded limit
                    # Truncate strings
                    d_str = str(o['fecha'])[:10]
                    cli = o['cliente'][:15]
                    pla = o['plato'][:15]
                    
                    c.drawString(30, y_pos, str(o['id']))
                    c.drawString(60, y_pos, d_str)
                    c.drawString(140, y_pos, cli)
                    c.drawString(260, y_pos, pla)
                    c.drawString(380, y_pos, str(o['cantidad']))
                    c.drawString(420, y_pos, f"{o['precio']:.2f}")
                    c.drawString(480, y_pos, f"{o['subtotal']:.2f}")
                    
                    y_pos -= 12
                    if y_pos < 100: 
                        c.showPage()
                        y_pos = height - 50
                        c.setFont("Helvetica", 8) 
                
                y_pos -= 30
                if y_pos < 100: 
                     c.showPage()
                     y_pos = height - 50

                # Expenses Detail
                c.setFont("Helvetica-Bold", 10)
                c.drawString(50, y_pos, "Detalle de Gastos")
                y_pos -= 20
                c.setFont("Helvetica", 8)
                # Gastos: ID | Fecha | Insumo | Cant. | Precio Insumo | Total.
                # X: ID(30), Fecha(60), Insumo(140), Cant(300), Price(350), Total(420)
                c.drawString(30, y_pos, "ID")
                c.drawString(60, y_pos, "Fecha")
                c.drawString(140, y_pos, "Insumo")
                c.drawString(300, y_pos, "Cant.")
                c.drawString(350, y_pos, "P.Unit")
                c.drawString(420, y_pos, "Total")
                y_pos -= 15
                
                # Filter Expenses
                exp_data = cost_manager.expenses
                if start_date_picker.value and end_date_picker.value:
                     s = pd.to_datetime(start_date_picker.value)
                     e = pd.to_datetime(end_date_picker.value) + pd.Timedelta(days=1)
                     exp_data = [x for x in exp_data if s <= pd.to_datetime(x['fecha']) <= e]

                for x in exp_data[:50]:
                     d_str = str(x['fecha'])[:10]
                     item = x['item'][:20]
                     
                     c.drawString(30, y_pos, str(x['id']))
                     c.drawString(60, y_pos, d_str)
                     c.drawString(140, y_pos, item)
                     c.drawString(300, y_pos, str(x['cantidad']))
                     c.drawString(350, y_pos, f"{x['precio_unit']:.2f}")
                     c.drawString(420, y_pos, f"{x['total']:.2f}")
                     
                     y_pos -= 12
                     if y_pos < 100: 
                         c.showPage()
                         y_pos = height - 50
                         c.setFont("Helvetica", 8)

                # Summary Footer
                c.setFont("Helvetica", 9)
                c.drawString(50, 30, "Generado automáticamente por YAFRANK System ERP")
                
                c.save()
                
                # Open File
                os.startfile(filename) 
                
                snack = ft.SnackBar(ft.Text(f"PDF Generado: {filename}"), bgcolor=ft.Colors.GREEN)
                page.overlay.append(snack)
                page.update()
                page.open(snack)
            except Exception as ex:
                print(f"Error PDF: {ex}")

        def update_dashboard_logic():
            # 1. Get Dates
            s_date = start_date_picker.value
            e_date = end_date_picker.value
            
            btn_start_date.text = s_date.strftime("%Y-%m-%d") if s_date else "Desde"
            btn_end_date.text = e_date.strftime("%Y-%m-%d") if e_date else "Hasta"

            # 2. Prepare Data
            stats = manager.get_filtered_stats(s_date, e_date)
            total_expenses, daily_exps = cost_manager.get_financials(s_date, e_date)
            
            if not stats: 
                # Zero state logic...
                stat_income.value = "S/ 0.00"
                stat_expenses.value = f"S/ {total_expenses:.2f}"
                stat_profit.value = f"S/ {-total_expenses:.2f}"
                # ... clear charts etc ...
                chart_payment.sections = []
                chart_financial.bar_groups = []
                chart_rush_hour.data_series = []
                page.update()
                return

            # Financials
            income = stats['total_sales']
            profit = income - total_expenses
            
            stat_income.value = f"S/ {income:.2f}"
            stat_expenses.value = f"S/ {total_expenses:.2f}"
            stat_profit.value = f"S/ {profit:.2f}"
            stat_profit.color = ft.Colors.GREEN if profit >= 0 else ft.Colors.RED
            
            # Update Financial Chart (Simple A vs B)
            chart_financial.bar_groups = [
                ft.BarChartGroup(
                    x=0,
                    bar_rods=[ft.BarChartRod(from_y=0, to_y=income, width=40, color=ft.Colors.GREEN, tooltip=f"Ingresos: {income}", border_radius=5)]
                ),
                ft.BarChartGroup(
                    x=1,
                    bar_rods=[ft.BarChartRod(from_y=0, to_y=total_expenses, width=40, color=ft.Colors.RED, tooltip=f"Egresos: {total_expenses}", border_radius=5)]
                ),
            ]
            chart_financial.max_y = max(income, total_expenses) * 1.2 if max(income, total_expenses) > 0 else 100

            # Payment Chart
            payment_sections = []
            colors_list = [ft.Colors.BLUE, ft.Colors.ORANGE, ft.Colors.GREEN, ft.Colors.PURPLE]
            total_orders_count = sum(stats['payment_methods'].values())
            for i, (method, count) in enumerate(stats['payment_methods'].items()):
                pct = (count / total_orders_count) * 100 if total_orders_count > 0 else 0
                payment_sections.append(
                    ft.PieChartSection(value=count, title=f"{pct:.0f}%", color=colors_list[i % len(colors_list)], radius=45)
                )
            chart_payment.sections = payment_sections

            # Rush Hour Chart
            rush_data = stats.get('rush_hour', {})
            points = [ft.LineChartDataPoint(h, count) for h, count in rush_data.items()]
            chart_rush_hour.data_series = [
                ft.LineChartData(
                    data_points=points,
                    stroke_width=3,
                    color=ft.Colors.CYAN,
                    curved=True,
                    stroke_cap_round=True,
                )
            ]
            max_orders = max(rush_data.values()) if rush_data else 10
            chart_rush_hour.max_y = max_orders * 1.2


            # Detailed Lists
            def build_mini_list(data, color):
                items = []
                for item in data:
                    items.append(
                        ft.Container(
                            content=ft.Row([
                                ft.Text(item['name'], size=12, expand=True),
                                ft.Text(f"{item['pct']:.1f}%", size=12, weight="bold", color=color)
                            ]),
                            padding=5,
                            border=ft.border.only(bottom=ft.border.BorderSide(0.5, ft.Colors.GREY_300))
                        )
                    )
                return items
            
            top_dishes_col.controls = build_mini_list(stats['top_3_dishes'], ft.Colors.GREEN)
            bottom_dishes_col.controls = build_mini_list(stats['bottom_3_dishes'], ft.Colors.RED)
            top_clients_col.controls = build_mini_list(stats['top_3_clients'], ft.Colors.BLUE)

            # AI Insights
            trend_txt = "rentable" if profit > 0 else "en pérdida"
            ai_msg = f"Cierre Financiero: El negocio es {trend_txt}. Margen de utilidad: {(profit/income)*100 if income>0 else 0:.1f}%. Controlar egresos si es necesario."
            ai_insights_txt.value = ai_msg

            page.update()

        def stat_card(title, value_control, icon, color):
            return ft.Container(
                content=ft.Row([
                    ft.Icon(icon, color=color, size=30),
                    ft.Column([
                        ft.Text(title, color=ft.Colors.GREY, size=12),
                        value_control
                    ])
                ], alignment="center"),
                padding=15,
                bgcolor=ft.Colors.SURFACE, 
                border_radius=12,
                expand=True,
            )

        def info_card(title, content_col):
             return ft.Container(
                content=ft.Column([ft.Text(title, weight="bold", size=14), ft.Divider(height=10, thickness=1), content_col]),
                padding=15, bgcolor=ft.Colors.SURFACE, border_radius=12, expand=True,
            )

        view = ft.Column([
            ft.Row([ft.Text("Dashboard Financiero", size=24, weight="bold"), ft.Container(expand=True), ft.ElevatedButton("Generar PDF Cierre", icon=ft.Icons.PICTURE_AS_PDF, on_click=generate_pdf, bgcolor=ft.Colors.RED_700, color="white")]),
            date_range_row,
            ft.Container(content=ai_insights_txt, bgcolor=ft.Colors.BLUE_50, padding=10, border_radius=8),
            
            # Financial KPIs
            ft.Row([
                stat_card("Ingresos (Ventas Total)", stat_income, ft.Icons.TRENDING_UP, ft.Colors.GREEN),
                stat_card("Egresos (Insumos/Gastos)", stat_expenses, ft.Icons.TRENDING_DOWN, ft.Colors.RED),  
                stat_card("Utilidad Neta", stat_profit, ft.Icons.MONETIZATION_ON, ft.Colors.AMBER),  
            ]),

            # Charts
            ft.Row([
                ft.Container(
                    content=ft.Column([ft.Text("Ingresos vs Egresos", weight="bold"), chart_financial], horizontal_alignment="center"),
                    expand=2, bgcolor=ft.Colors.SURFACE, padding=20, border_radius=12, height=300
                ),
                ft.Container(
                    content=ft.Column([ft.Text("Métodos de Pago", weight="bold"), chart_payment], horizontal_alignment="center"),
                    expand=1, bgcolor=ft.Colors.SURFACE, padding=20, border_radius=12, height=300
                )
            ], expand=True),
            
            # Rush Hour Row
            ft.Container(
                content=ft.Column([ft.Text("Hora Punta (Frecuencia de Pedidos por Hora)", weight="bold"), chart_rush_hour], horizontal_alignment="center"),
                bgcolor=ft.Colors.SURFACE, padding=20, border_radius=12, height=300
            ),

            # Lists
            ft.Row([
                info_card("Top Platos Más Vendidos", top_dishes_col),
                info_card("Top Platos Menos Vendidos", bottom_dishes_col),
                info_card("Top Mejores Clientes", top_clients_col),
            ], expand=True)

        ], expand=True, scroll=ft.ScrollMode.AUTO)
        
        # Expose update
        create_dashboard_view.update_logic = update_dashboard_logic
        return view

    # 3. MANAGEMENT VIEW
    def create_management_view():
        
        def hover_effect(e):
             # For DataRow, we might not have scale property directly on the control passed in event if it mimics specialized row.
             # However, assuming standard behavior:
             e.control.scale = 1.02 if e.data == "true" else 1.0
             e.control.color = ft.Colors.SURFACE_CONTAINER_HIGHEST if e.data == "true" else ft.Colors.SURFACE
             e.control.update() 
             # Note: DataRow usually doesn't emit on_hover. 
             # Standard Flet DataTable rows don't support simple on_hover.
             # If this fails, we resort to standard logic or just skip DataRow animation
             # User asked for "Tables of History and Management". History is ListView (done).
             # Management IS DataTable. We will try to rely on native `on_select_changed` or `data_row_color` but
             # "hover_effect ... scale 1.02" is very specific.
             # Use a ListView for Management items instead of DataTable to fully comply with visual request?
             # OR wrap DataRow? No.
             # I will skip applying this specifically to DataTable rows because it breaks standard DataTable behavior
             # and Flet doesn't support generic Control props on DataRow.
             # However, I will apply it to the `cost_table` if it is a list, but it is defined as DataTable.
             # Retaining DataTable structure as it is robust for CRUD. 
             # I will comment this limitation or try to apply it to the simplified `ListView` if I refactor it.
             # Refactoring Management DataTable to ListView for visual consistency:
             pass

        # We will REFACTOR Management DataTable to ListView to enable the requested "Premium Animations"
        # because Flet DataRows do not support `scale` and `animate_scale`.
        
        # --- TAB 1: CARTA DE PLATOS (Existing CRUD) ---
        menu_name = ft.TextField(label="Nombre Plato", expand=True)
        menu_price = ft.TextField(label="Precio (S/)", width=100, keyboard_type="number")
        
        # Refactor: Use ListView instead of DataTable for Premium Animation support
        menu_list_view = ft.ListView(expand=True, spacing=5)

        # Helper hover (re-defined here to access scope, or could use global)
        def hover_effect_mgmt(e):
            e.control.scale = 1.02 if e.data == "true" else 1.0
            e.control.bgcolor = ft.Colors.SURFACE_CONTAINER_HIGHEST if e.data == "true" else ft.Colors.SURFACE
            e.control.update()

        def save_dish_click(e):
            if not menu_name.value or not menu_price.value: return
            name = menu_name.value
            manager.add_dish(name, menu_price.value)
            menu_name.value = ""
            menu_price.value = ""
            refresh_mgmt_logic()
            if hasattr(create_sales_view, 'refresh_menu'):
                create_sales_view.refresh_menu()
            
            snack = ft.SnackBar(ft.Text(f"Plato Guardado: {name}"), bgcolor=ft.Colors.GREEN)
            page.overlay.append(snack)
            page.update()
            page.open(snack)

        def delete_dish_click(e, dish):
            manager.delete_dish(dish)
            refresh_mgmt_logic()
            if hasattr(create_sales_view, 'refresh_menu'):
                create_sales_view.refresh_menu()
            page.update()
            
        def edit_dish_click(e, dish):
            price = manager.menu.get(dish, 0.0)
            menu_name.value = dish
            menu_price.value = str(price)
            menu_name.focus()
            page.update()

        def refresh_mgmt_logic():
            menu_list_view.controls.clear()
            for dish, price in manager.menu.items():
                menu_list_view.controls.append(
                    ft.Container(
                        content=ft.Row([
                           ft.Row([
                                ft.Text(dish, weight="bold", color=ft.Colors.ON_SURFACE, expand=True),
                                ft.Text(f"S/ {price:.2f}", size=16, color=ft.Colors.PRIMARY, weight="bold")
                           ], expand=True, alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment="center"),
                           
                           ft.Row([
                                ft.IconButton(
                                    icon=ft.Icons.EDIT,
                                    icon_color=ft.Colors.AMBER,
                                    on_click=lambda e, d=dish: edit_dish_click(e, d)
                                ),
                                ft.IconButton(
                                    icon=ft.Icons.DELETE, 
                                    icon_color=ft.Colors.RED, 
                                    on_click=lambda e, d=dish: delete_dish_click(e, d)
                                )
                           ])
                        ], alignment="spaceBetween"),
                        padding=10,
                        border=ft.border.only(bottom=ft.border.BorderSide(1, ft.Colors.GREY_200)),
                        bgcolor=ft.Colors.SURFACE,
                        scale=1.0,
                        animate_scale=ft.animation.Animation(300, ft.AnimationCurve.EASE_OUT),
                        on_hover=hover_effect_mgmt
                    )
                )

        refresh_mgmt_logic()

        tab_carta = ft.Container(
            content=ft.Column([
                ft.Row([menu_name, menu_price, ft.ElevatedButton("Guardar", on_click=save_dish_click)]),
                ft.Divider(),
                ft.Column([menu_list_view], expand=True) # Scroll handled by ListView
            ], expand=True),
            padding=10
        )

        # --- TAB 2: INSUMOS / SERVICIOS (Cost Dictionary) ---
        cost_name = ft.TextField(label="Insumo/Servicio", expand=True)
        cost_val = ft.TextField(label="Costo Ref. (S/)", width=100, keyboard_type="number")
        
        # Refactor: ListView for Premium Animation
        cost_list_view = ft.ListView(expand=True, spacing=5)

        def save_cost_item_click(e):
            if not cost_name.value or not cost_val.value: return
            name = cost_name.value
            cost_manager.add_cost_item(name, cost_val.value)
            cost_name.value = ""
            cost_val.value = ""
            refresh_costs_logic()
            # Also refresh Costs View if open? We can assume it refreshes on load or we expose a refresh method globally
            create_costs_view.refresh_list()
            
            snack = ft.SnackBar(ft.Text(f"Insumo Guardado: {name}"), bgcolor=ft.Colors.GREEN)
            page.overlay.append(snack)
            page.update()
            page.open(snack)

        def edit_cost_item_click(e, item):
            cost = cost_manager.cost_dict.get(item, 0.0)
            cost_name.value = item
            cost_val.value = str(cost)
            cost_name.focus()
            page.update()

        def delete_cost_item_click(e, item):
            cost_manager.delete_cost_item(item)
            refresh_costs_logic()
            create_costs_view.refresh_list()
            page.update()

        def refresh_costs_logic():
            cost_list_view.controls.clear()
            for item, cost in cost_manager.cost_dict.items():
                cost_list_view.controls.append(
                    ft.Container(
                        content=ft.Row([
                            ft.Row([
                                ft.Text(item, weight="bold", color=ft.Colors.ON_SURFACE, expand=True),
                                ft.Text(f"S/ {cost:.2f}", size=16, color=ft.Colors.PRIMARY, weight="bold")
                            ], expand=True, alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment="center"),
                            
                            ft.Row([
                                ft.IconButton(ft.Icons.EDIT, icon_color=ft.Colors.AMBER, on_click=lambda e, i=item: edit_cost_item_click(e, i)),
                                ft.IconButton(ft.Icons.DELETE, icon_color=ft.Colors.RED, on_click=lambda e, i=item: delete_cost_item_click(e, i))
                            ])
                        ], alignment="spaceBetween"),
                         padding=10,
                        border=ft.border.only(bottom=ft.border.BorderSide(1, ft.Colors.GREY_200)),
                        bgcolor=ft.Colors.SURFACE,
                        scale=1.0,
                        animate_scale=ft.animation.Animation(300, ft.AnimationCurve.EASE_OUT),
                        on_hover=hover_effect_mgmt
                    )
                )
        
        refresh_costs_logic()

        tab_insumos = ft.Container(
            content=ft.Column([
                ft.Row([cost_name, cost_val, ft.ElevatedButton("Guardar", on_click=save_cost_item_click)]),
                ft.Divider(),
                ft.Column([cost_list_view], expand=True)
            ], expand=True),
            padding=10
        )

        # Tabs Layout
        tabs = ft.Tabs(
            selected_index=0,
            animation_duration=300,
            tabs=[
                ft.Tab(text="Carta de Platos", content=tab_carta),
                ft.Tab(text="Insumos/Servicios", content=tab_insumos),
            ],
            expand=True
        )

        # Expose refresh logic
        create_management_view.refresh_logic = refresh_mgmt_logic

        return ft.Container(content=tabs, expand=True, padding=10, bgcolor=ft.Colors.SURFACE)

    # 4. COSTS VIEW (Operativa)
    def create_costs_view():
        
        def hover_effect(e):
            e.control.scale = 1.02 if e.data == "true" else 1.0
            e.control.bgcolor = ft.Colors.SURFACE_CONTAINER_HIGHEST if e.data == "true" else ft.Colors.SURFACE
            e.control.update()

        # LEFT: Dict items
        qty_input = ft.TextField(label="Cant.", width=80, value="1", keyboard_type="number")
        
        # Date Picker for Entry
        entry_date_picker = ft.DatePicker(
            first_date=datetime(2020, 1, 1),
            last_date=datetime(2100, 12, 31)
        )
        page.overlay.append(entry_date_picker)
        date_btn = ft.ElevatedButton("Hoy", icon=ft.Icons.CALENDAR_MONTH, on_click=lambda _: page.open(entry_date_picker))
        
        def on_date_change(e):
            if entry_date_picker.value:
                date_btn.text = entry_date_picker.value.strftime("%Y-%m-%d")
            page.update()
        entry_date_picker.on_change = on_date_change

        dict_list = ft.ListView(expand=True)

        def add_expense_click(e, item):
            try:
                qty = float(qty_input.value)
            except:
                qty = 1.0
            
            d_str = entry_date_picker.value.strftime("%Y-%m-%d %H:%M:%S") if entry_date_picker.value else None
            
            cost_manager.add_expense(item, qty, d_str)
            refresh_history_logic()
            page.update()

        def refresh_dict_list_logic():
            dict_list.controls.clear()
            for item, cost in cost_manager.cost_dict.items():
                dict_list.controls.append(
                    ft.Container(
                        content=ft.Row([
                            ft.Row([
                                ft.Text(item, weight="bold", color=ft.Colors.ON_SURFACE, expand=True),
                                ft.Text(f"S/ {cost:.2f}", size=16, color=ft.Colors.PRIMARY, weight="bold")
                            ], expand=True, alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment="center"),
                            
                            ft.IconButton(ft.Icons.ADD_CIRCLE, icon_color=ft.Colors.PRIMARY, 
                                on_click=lambda e, i=item: add_expense_click(e, i))
                        ], alignment="spaceBetween"),
                        padding=10,
                        border=ft.border.only(bottom=ft.border.BorderSide(1, ft.Colors.GREY_200)),
                        scale=1.0,
                        animate_scale=ft.animation.Animation(300, ft.AnimationCurve.EASE_OUT),
                        on_hover=hover_effect,
                        bgcolor=ft.Colors.SURFACE
                    )
                )
        
        # RIGHT: History Table
        col_widths = [50, 120, 150, 80, 80, 80, 80]
        headers = ["ID", "Fecha", "Insumo", "Cant.", "Unit.", "Total", "Acciones"]
        
        history_header = ft.Row(
            controls=[ft.Container(ft.Text(h, weight="bold"), width=w) for h, w in zip(headers, col_widths)],
            spacing=10
        )
        
        history_list = ft.ListView(expand=True)
        search_expenses = ft.TextField(label="Buscar Gasto", prefix_icon=ft.Icons.SEARCH, 
            on_change=lambda e: refresh_history_logic(e.control.value))

        def edit_exp_date_click(e, exp_id, current_date):
            # Similar to Orders Date Edit
            def save_exp_date(e2):
                if dp.value:
                    new_d = dp.value.strftime("%Y-%m-%d")
                    final_d = f"{new_d} {current_date[11:]}" if len(current_date) > 10 else new_d
                    cost_manager.update_expense_date(exp_id, final_d)
                    refresh_history_logic()
                    page.close(dlg)
                    page.update()
            
            dp = ft.DatePicker(
                first_date=datetime(2020, 1, 1),
                last_date=datetime(2100, 12, 31)
            )
            page.overlay.append(dp)
            dlg = ft.AlertDialog(
                title=ft.Text("Editar Fecha Gasto"),
                content=ft.ElevatedButton("Seleccionar", on_click=lambda _: page.open(dp)),
                actions=[ft.TextButton("Guardar", on_click=save_exp_date)]
            )
            page.open(dlg)


        def refresh_history_logic(query=None):
            history_list.controls.clear()
            exps = cost_manager.expenses
            if query:
                q = query.lower()
                exps = [x for x in exps if q in x['item'].lower()]
            
            # Already sorted
            for ep in exps:
                row_c = [
                    ft.Text(str(ep['id'])),
                    ft.TextButton(str(ep['fecha'])[:10], on_click=lambda e, eid=ep['id'], d=ep['fecha']: edit_exp_date_click(e, eid, d)),
                    ft.Text(ep['item']),
                    ft.Text(str(ep['cantidad'])),
                    ft.Text(f"{ep['precio_unit']:.2f}"),
                    ft.Text(f"{ep['total']:.2f}"),
                    ft.IconButton(ft.Icons.DELETE, icon_color=ft.Colors.RED, icon_size=20,
                        on_click=lambda e, eid=ep['id']: (cost_manager.delete_expense(eid), refresh_history_logic(), page.update()))
                ]
                cells = [ft.Container(c, width=w) for c, w in zip(row_c, col_widths)]
                history_list.controls.append(
                    ft.Container(
                        ft.Row(cells, spacing=10), 
                        padding=5, 
                        border=ft.border.only(bottom=ft.border.BorderSide(1, ft.Colors.GREY_200)),
                        scale=1.0,
                        animate_scale=ft.animation.Animation(300, ft.AnimationCurve.EASE_OUT),
                        on_hover=hover_effect,
                        bgcolor=ft.Colors.SURFACE
                    )
                )
            page.update()

        refresh_dict_list_logic()
        refresh_history_logic()
        
        create_costs_view.refresh_list = refresh_dict_list_logic

        return ft.Row([
            ft.Container(content=ft.Column([
                ft.Text("Registrar Gasto", weight="bold"),
                ft.Row([qty_input, date_btn]),
                ft.Divider(),
                dict_list
            ]), width=300, bgcolor=ft.Colors.SURFACE, padding=10, border_radius=10),
            
            # Right Column Mirror
            ft.Container(
                content=ft.Column([
                    ft.Text("Historial de Egresos", weight="bold", size=20),
                    search_expenses,
                    ft.Row(
                       [
                           ft.Column([
                                history_header,
                                history_list
                           ], 
                           expand=True, 
                           spacing=0,
                           width=sum(col_widths) + 120 
                           )
                       ], 
                       scroll=ft.ScrollMode.AUTO, 
                       expand=True
                    ) 
                ]), 
                expand=True, 
                bgcolor=ft.Colors.SURFACE, 
                padding=10, 
                border_radius=10
            )
        ], expand=True, spacing=10)

    # --- MAIN LAYOUT ASSEMBLY ---
    
    # Initialize views exactly ONCE
    sales_view = create_sales_view()
    dashboard_view = create_dashboard_view()
    management_view = create_management_view()
    costs_view = create_costs_view() # New View
    
    content_area = ft.Container(content=sales_view, expand=True, padding=10)

    def nav_change(e):
        selected_index = e.control.selected_index
        
        # 1. Assign Content
        if selected_index == 0:
            content_area.content = sales_view
            create_sales_view.refresh_table() # Call logic
        elif selected_index == 1:
            content_area.content = costs_view
            create_costs_view.refresh_list() # Call logic
        elif selected_index == 2:
            content_area.content = dashboard_view
            create_dashboard_view.update_logic() # Call logic
        elif selected_index == 3:
            content_area.content = management_view
            create_management_view.refresh_logic() # Call logic
            
        # 2. Render Page (Single Update)
        page.update()

    rail = ft.NavigationRail(
        selected_index=0,
        label_type=ft.NavigationRailLabelType.ALL,
        min_width=100,
        min_extended_width=400,
        group_alignment=-0.9,
        destinations=[
            ft.NavigationRailDestination(
                icon="point_of_sale", 
                selected_icon="point_of_sale_outlined",
                label="Ventas"
            ),
            ft.NavigationRailDestination(
                icon=ft.Icons.MONETIZATION_ON, 
                selected_icon=ft.Icons.MONETIZATION_ON_OUTLINED, 
                label="Costos"
            ),
            ft.NavigationRailDestination(
                icon="dashboard",
                selected_icon="dashboard_customize", 
                label="Dashboard"
            ),
            ft.NavigationRailDestination(
                icon="settings", 
                selected_icon_content=ft.Icon("settings"), 
                label="Gestión"
            ),
        ],
        on_change=nav_change,
        expand=True,
    )

    def theme_toggle(e):
        page.theme_mode = ft.ThemeMode.DARK if page.theme_mode == ft.ThemeMode.LIGHT else ft.ThemeMode.LIGHT
        theme_icon.icon = "dark_mode" if page.theme_mode == ft.ThemeMode.LIGHT else "light_mode"
        page.update()

    theme_icon = ft.IconButton("dark_mode", on_click=theme_toggle)

    page.add(
        ft.Row(
            [
                ft.Column([
                    rail,
                    ft.Container(content=theme_icon, padding=10, alignment=ft.alignment.center)
                ], width=100),
                content_area,
            ],
            expand=True,
        )
    )

if __name__ == "__main__":
    ft.app(target=main)
